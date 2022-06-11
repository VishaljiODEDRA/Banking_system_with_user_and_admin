import sys
import subprocess

subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
import tkinter
import tkinter.messagebox
from openpyxl import load_workbook


class BankingSystem:
    def __init__(self):
        # Do not add any parameter to this method.
        # Delete "pass" after adding code into this method.
        self.user_file()

    def user_file(self):
        self.wb = load_workbook(
            filename=
            "/Users/vishaljiodedra/OneDrive - Birmingham City University/CMP4266/base code - banking_system/User_info.xlsx"
        )
        self.ws = self.wb.active

    def save_user_file(self):
        self.wb.save(
            filename=
            "/Users/vishaljiodedra/OneDrive - Birmingham City University/CMP4266/base code - banking_system/User_info.xlsx"
        )

    def login(self):
        self.a = input("Enter your username: ")
        self.b = input("Enter your password: ")
        self.user_file()
        for row in range(1, 20):
            if self.a == str(self.ws.cell(row, 1).value) and self.b == str(
                    self.ws.cell(row, 2).value):
                User(
                    self.ws.cell(row, 1).value,
                    self.ws.cell(row, 2).value,
                    self.ws.cell(row, 3).value)
                if self.ws.cell(row, 3).value == "Admin":
                    admin = Admin(
                        self.ws.cell(row, 1).value,
                        self.ws.cell(row, 2).value,
                        self.ws.cell(row, 3).value)
                    print(
                        f'You are {self.ws.cell(row, 3).value} and logged in as {self.ws.cell(row, 1).value}.'
                    )
                    admin.display_Ad_menu()
                    break
                elif self.ws.cell(row, 3).value == "Customer":
                    customer = Customer(
                        self.ws.cell(row, 1).value,
                        self.ws.cell(row, 2).value,
                        self.ws.cell(row, 3).value,
                        self.ws.cell(row, 5).value)
                    print(
                        f'You are logged in as {self.ws.cell(row, 1).value}.')
                    customer.display_Cus_menu()
                    break

        else:
            print("Invalid username/password")
            self.login()

    def run_app(self):
        print("------ Welcome to V - Banking System ------")
        self.login()


class User(BankingSystem):
    def __init__(self, name, password, uType):
        self.name = name
        self.password = password
        self.uType = uType


class Admin(User):
    def __init__(self, name, password, uType):
        User.__init__(self, name, password, uType)
        self.adm_menu = ("Please select an option:", "  1 - Customer Summary",
                         "  2 - Financial Forecast",
                         "  3 - Transfer Money - GUI",
                         "  4 - Account management - GUI")
        self.menu = ("Name", "Address", "Account type", "Balance", "Overdraft",
                     "Interest rate")
        self.finan_menu = ("Name", "Total No of Acc.", "Tot. Money in Bank",
                           "Money after a year", "Total deposited",
                           "Total withdrawn", "Total transferred")

    def display_Ad_menu(self):
        for i in self.adm_menu:
            print(i)
        ask_input2 = input("Enter a number to select your option:")
        if ask_input2 == "1":
            self.Cust_Summary()
        elif ask_input2 == "2":
            self.Finan_Forecast()
        elif ask_input2 == "3":
            self.Trans_Money_GUI()
        elif ask_input2 == "4":
            self.Acc_Management_GUI()
        else:
            print("Invalid input...\nPlease enter again")
            self.display_Ad_menu()

    def Cust_Summary(self):
        self.user_file()
        acc_and_account_no = []
        print()
        for i in self.menu:
            print(f'{i:<20}', end='')
        print("\n")
        for row in range(1, 20):
            if self.ws.cell(row, 3).value == "Customer" and self.ws.cell(
                    row, 1).value + str(self.ws.cell(
                        row, 7).value) not in acc_and_account_no:
                print(
                    f'{self.ws.cell(row, 1).value:<20}{self.ws.cell(row, 5).value:<20}{self.ws.cell(row, 6).value:<20}£{self.ws.cell(row, 4).value:<20}£{self.ws.cell(row, 8).value:<20}{self.ws.cell(row, 9).value}%'
                )
                for x in range(1, 20):
                    if x != row and self.ws.cell(row, 1).value == self.ws.cell(
                            x, 1).value:
                        y = self.ws.cell(x, 1).value + str(
                            self.ws.cell(x, 7).value)
                        acc_and_account_no.append(y)
                        z = ""
                        print(
                            f'{z:<40}{self.ws.cell(x, 6).value:<20}£{self.ws.cell(x, 4).value:<20}£{self.ws.cell(x, 8).value:<20}{self.ws.cell(x, 9).value}%'
                        )

    def Finan_Forecast(self):
        self.user_file()
        acc_and_account_no = []
        acc = []
        print()
        for i in self.finan_menu:
            print(f'{i:<20}', end='')
        print("\n")
        for row in range(1, 20):
            tot_deposited = 0
            tot_withdrawn = 0
            Total_transferred = 0
            self.bal = 0
            self.acc_numb = 0
            if self.ws.cell(row, 3).value == "Customer" and self.ws.cell(
                    row, 1).value + str(self.ws.cell(
                        row, 7).value) not in acc_and_account_no:
                self.bal += self.ws.cell(row, 4).value
                intrest = (self.ws.cell(row, 4).value *
                           self.ws.cell(row, 9).value * 1) / 100
                bal_after_year = self.ws.cell(row, 4).value + intrest
                tot_deposited += self.ws.cell(row, 10).value
                tot_withdrawn += self.ws.cell(row, 11).value
                Total_transferred += self.ws.cell(row, 12).value
                self.acc_numb += 1
                for x in range(1, 20):
                    if x != row and self.ws.cell(row, 1).value == self.ws.cell(
                            x, 1).value:
                        y = self.ws.cell(x, 1).value + str(
                            self.ws.cell(x, 7).value)
                        acc_and_account_no.append(y)
                        self.bal += self.ws.cell(x, 4).value
                        intrest = (self.ws.cell(x, 4).value *
                                   self.ws.cell(x, 9).value * 1) / 100
                        bal_after_year = bal_after_year + self.ws.cell(
                            x, 4).value + intrest
                        tot_deposited += self.ws.cell(x, 10).value
                        tot_withdrawn += self.ws.cell(x, 11).value
                        Total_transferred += self.ws.cell(x, 12).value
                        self.acc_numb += 1
                        print(
                            f'{self.ws.cell(x, 1).value:<20}{self.acc_numb:<20}£{self.bal:<20}£{bal_after_year:<20}£{tot_deposited:<20}£{tot_withdrawn:<20}£{Total_transferred:<20}'
                        )
                        break
                else:
                    print(
                        f'{self.ws.cell(row, 1).value:<20}{self.acc_numb:<20}£{self.bal:<20}£{bal_after_year:<20}£{tot_deposited:<20}£{tot_withdrawn:<20}£{Total_transferred:<20}'
                    )

    def Trans_Money_GUI(self):
        self.mw = tkinter.Tk()
        self.mw.title("Money Transction Portal")

        self.top_frame1 = tkinter.Frame(self.mw)
        self.top_frame2 = tkinter.Frame(self.mw)
        self.top_frame3 = tkinter.Frame(self.mw)
        self.mid_frame1 = tkinter.Frame(self.mw)
        self.mid_frame2 = tkinter.Frame(self.mw)
        self.mid_frame3 = tkinter.Frame(self.mw)
        self.bottom_frame1 = tkinter.Frame(self.mw)
        self.bottom_frame2 = tkinter.Frame(self.mw)
        self.bottom_frame3 = tkinter.Frame(self.mw)
        self.bottom_frame4 = tkinter.Frame(self.mw)

        self.from_sms1 = tkinter.Label(
            self.top_frame1,
            text="Enter details, FROM which customer account:")
        self.from_sms2 = tkinter.Label(self.top_frame2, text="Username:")
        self.ask_from_username = tkinter.Entry(self.top_frame2, width=15)
        self.from_sms3 = tkinter.Label(self.top_frame3, text="Account Number:")
        self.ask_from_account = tkinter.Entry(self.top_frame3, width=15)

        self.to_sms1 = tkinter.Label(
            self.mid_frame1,
            text="\nEnter details, TO which customer account:")
        self.to_sms2 = tkinter.Label(self.mid_frame2, text="Username:")
        self.ask_to_username = tkinter.Entry(self.mid_frame2, width=15)
        self.to_sms3 = tkinter.Label(self.mid_frame3, text="Account Number:")
        self.ask_to_account = tkinter.Entry(self.mid_frame3, width=15)

        self.blank_line = tkinter.Label(self.bottom_frame1, text=" ")
        self.amount_sms = tkinter.Label(self.bottom_frame2,
                                        text="ENTER AMOUNT:")
        self.ask_amount = tkinter.Entry(self.bottom_frame2, width=15)
        self.make_transfer = tkinter.Button(self.bottom_frame3,
                                            text="MAKE TRANSACTION",
                                            command=self.Money_transfer)
        self.money_tran_reset = tkinter.Button(
            self.bottom_frame3,
            text="Reset",
            command=self.Money_transfer_reset)
        self.exit_button2 = tkinter.Button(self.bottom_frame4,
                                           text="Exit",
                                           command=self.mw.destroy)

        self.from_sms1.pack(side="left")
        self.from_sms2.pack(side="left")
        self.ask_from_username.pack(side="left")
        self.from_sms3.pack(side="left")
        self.ask_from_account.pack(side="left")

        self.to_sms1.pack(side="left")
        self.to_sms2.pack(side="left")
        self.ask_to_username.pack(side="left")
        self.to_sms3.pack(side="left")
        self.ask_to_account.pack(side="left")

        self.blank_line.pack(side="left")
        self.amount_sms.pack(side="left")
        self.ask_amount.pack(side="left")
        self.make_transfer.pack(side="left")
        self.money_tran_reset.pack(side="left")
        self.exit_button2.pack(side="left")

        self.top_frame1.pack()
        self.top_frame2.pack()
        self.top_frame3.pack()
        self.mid_frame1.pack()
        self.mid_frame2.pack()
        self.mid_frame3.pack()
        self.bottom_frame1.pack()
        self.bottom_frame2.pack()
        self.bottom_frame3.pack()
        self.bottom_frame4.pack()

        tkinter.mainloop()

    def Acc_Management_GUI(self):
        self.mw = tkinter.Tk()
        self.mw.title("Account Management Portal")

        self.top_f1 = tkinter.Frame(self.mw)
        self.top_f2 = tkinter.Frame(self.mw)
        self.top_f3 = tkinter.Frame(self.mw)
        self.mid_f1 = tkinter.Frame(self.mw)
        self.mid_f2 = tkinter.Frame(self.mw)
        self.mid_f3 = tkinter.Frame(self.mw)
        self.bottom_f1 = tkinter.Frame(self.mw)
        self.bottom_f2 = tkinter.Frame(self.mw)
        self.bottom_f3 = tkinter.Frame(self.mw)
        self.bottom_f4 = tkinter.Frame(self.mw)
        self.bottom_f5 = tkinter.Frame(self.mw)
        self.bottom_f6 = tkinter.Frame(self.mw)
        self.bottom_f7 = tkinter.Frame(self.mw)
        self.bottom_f8 = tkinter.Frame(self.mw)
        self.bottom_f9 = tkinter.Frame(self.mw)
        self.bottom_f10 = tkinter.Frame(self.mw)
        self.bottom_f11 = tkinter.Frame(self.mw)

        self.heading = tkinter.Label(
            self.top_f1, text="Enter information to Add an account: ")
        self.e_username = tkinter.Label(self.top_f2, text="Username: ")
        self.e_username1 = tkinter.Entry(self.top_f2, width=15)
        self.e_password = tkinter.Label(self.top_f3, text="Password: ")
        self.e_password1 = tkinter.Entry(self.top_f3, width=15)
        self.e_user_type = tkinter.Label(self.mid_f1, text="User type: ")
        self.e_user_type1 = tkinter.Entry(self.mid_f1, width=15)
        self.e_Balance = tkinter.Label(self.mid_f2, text="Balance: ")
        self.e_Balance1 = tkinter.Entry(self.mid_f2, width=15)
        self.e_Address = tkinter.Label(self.mid_f3, text="Address: ")
        self.e_Address1 = tkinter.Entry(self.mid_f3, width=15)
        self.e_acc_type = tkinter.Label(self.bottom_f1, text="Account type: ")
        self.e_acc_type1 = tkinter.Entry(self.bottom_f1, width=15)
        self.e_acc_no = tkinter.Label(self.bottom_f2, text="Account Number: ")
        self.e_acc_no1 = tkinter.Entry(self.bottom_f2, width=15)
        self.e_overdraft = tkinter.Label(self.bottom_f3,
                                         text="Overdraft limit: ")
        self.e_overdraft1 = tkinter.Entry(self.bottom_f3, width=15)
        self.e_intrest = tkinter.Label(self.bottom_f4,
                                       text="Rate of Intrest: ")
        self.e_intrest1 = tkinter.Entry(self.bottom_f4, width=15)
        self.add_button = tkinter.Button(self.bottom_f5,
                                         text="Add Account",
                                         command=self.add_account)
        self.add_reset_button = tkinter.Button(self.bottom_f5,
                                               text="Reset",
                                               command=self.add_reset_form)
        self.heading2 = tkinter.Label(
            self.bottom_f6, text="\nEnter information to Delete an account: ")
        self.d_username = tkinter.Label(self.bottom_f7, text="Username: ")
        self.d_username1 = tkinter.Entry(self.bottom_f7, width=15)
        self.d_password = tkinter.Label(self.bottom_f8, text="Password: ")
        self.d_password1 = tkinter.Entry(self.bottom_f8, width=15)
        self.d_acc_no = tkinter.Label(self.bottom_f9, text="Account Number: ")
        self.d_acc_no1 = tkinter.Entry(self.bottom_f9, width=15)

        self.delete_button = tkinter.Button(self.bottom_f10,
                                            text="Delete Account",
                                            command=self.delete_account)
        self.delete_reset_button = tkinter.Button(
            self.bottom_f10, text="Reset", command=self.delete_reset_form)
        self.exit_button1 = tkinter.Button(self.bottom_f11,
                                           text="Exit",
                                           command=self.mw.destroy)

        self.heading.pack(side="left")
        self.e_username.pack(side="left")
        self.e_username1.pack(side="left")
        self.e_password.pack(side="left")
        self.e_password1.pack(side="left")

        self.e_user_type.pack(side="left")
        self.e_user_type1.pack(side="left")
        self.e_Balance.pack(side="left")
        self.e_Balance1.pack(side="left")
        self.e_Address.pack(side="left")

        self.e_Address1.pack(side="left")
        self.e_acc_type.pack(side="left")
        self.e_acc_type1.pack(side="left")
        self.e_acc_no.pack(side="left")
        self.e_acc_no1.pack(side="left")
        self.e_overdraft.pack(side="left")
        self.e_overdraft1.pack(side="left")
        self.e_intrest.pack(side="left")
        self.e_intrest1.pack(side="left")

        self.add_button.pack(side="left")
        self.delete_button.pack(side="left")
        self.add_reset_button.pack(side="left")
        self.heading2.pack(side="left")
        self.d_username.pack(side="left")
        self.d_username1.pack(side="left")
        self.d_password.pack(side="left")
        self.d_password1.pack(side="left")
        self.d_acc_no.pack(side="left")
        self.d_acc_no1.pack(side="left")
        self.delete_button.pack(side="left")
        self.delete_reset_button.pack(side="left")
        self.exit_button1.pack(side="left")

        self.top_f1.pack()
        self.top_f2.pack()
        self.top_f3.pack()
        self.mid_f1.pack()
        self.mid_f2.pack()
        self.mid_f3.pack()
        self.bottom_f1.pack()
        self.bottom_f2.pack()
        self.bottom_f3.pack()
        self.bottom_f4.pack()
        self.bottom_f5.pack()
        self.bottom_f6.pack()
        self.bottom_f7.pack()
        self.bottom_f8.pack()
        self.bottom_f9.pack()
        self.bottom_f10.pack()
        self.bottom_f11.pack()

        tkinter.mainloop()

    def Money_transfer(self):
        self.user_file()
        try:
            if int(self.ask_amount.get()) > 0:
                for i in range(1, 20):
                    if self.ask_from_username.get() == self.ws.cell(
                            i, 1).value and self.ask_from_account.get() == str(
                                self.ws.cell(i, 7).value):
                        if int(self.ask_amount.get()) <= self.ws.cell(
                                i, 4).value and self.ws.cell(i, 4).value - int(
                                    self.ask_amount.get()) >= 0:
                            self.ws.cell(i,
                                         4).value -= int(self.ask_amount.get())
                            self.ws.cell(i, 12).value += int(
                                self.ask_amount.get())
                            for j in range(1, 20):
                                if j != i:
                                    if self.ask_to_username.get(
                                    ) == self.ws.cell(
                                            j, 1
                                    ).value and self.ask_to_account.get(
                                    ) == str(self.ws.cell(j, 7).value):
                                        self.ws.cell(j, 4).value += int(
                                            self.ask_amount.get())
                                        self.ws.cell(j, 12).value += int(
                                            self.ask_amount.get())
                                        tkinter.messagebox.showinfo(
                                            "TRANSACTION:",
                                            f'You have made £{self.ask_amount.get()} transfer from {self.ask_from_username.get()} account {self.ask_from_account.get()} to {self.ask_to_username.get()} account {self.ask_to_account.get()}.'
                                        )
                                        self.save_user_file()
                                        self.Money_transfer_reset()
                                        self.mw.destroy()
                                        break
                            else:
                                tkinter.messagebox.showinfo(
                                    "TRANSACTION:",
                                    f'Your Username/Password or Account Number  is wrong in "TO" section. please enter details again.'
                                )
                                self.ask_to_username.delete(0, 'end')
                                self.ask_to_account.delete(0, 'end')
                        else:
                            tkinter.messagebox.showinfo(
                                "TRANSACTION:",
                                f'The account with username {self.ask_from_username.get()} do not have sufficient balance to make transaction.'
                            )
                        break

                else:
                    tkinter.messagebox.showinfo(
                        "TRANSACTION:",
                        f'Your Username or Account Number is wrong in "FROM" section. please enter details again.'
                    )
                    self.ask_from_username.delete(0, 'end')
                    self.ask_from_account.delete(0, 'end')
            else:
                tkinter.messagebox.showinfo(
                    "TRANSACTION:",
                    f'The transaction amount can not be 0 or less than 0. Please enter amount again.'
                )
                self.ask_amount.delete(0, 'end')

        except ValueError:
            tkinter.messagebox.showinfo(
                "TRANSACTION:",
                f'The transaction amount can not be empty or any word or characters. Please enter proper amount.'
            )

    def add_account(self):
        self.user_file()
        try:
            for row in range(1, 20):
                if self.ws.cell(row, 1).value == None:
                    for i in range(1, 20):
                        if self.e_username1.get() == self.ws.cell(
                                i, 1).value and self.e_acc_type1.get(
                                ) == "Current Account" == self.ws.cell(
                                    i, 6).value:
                            tkinter.messagebox.showinfo(
                                "Account Management:",
                                f'The user {self.e_username1.get()} already has Current Account with same username. As per policy no more then one Current Account with same username.'
                            )
                            self.add_reset_form()
                            break
                    else:
                        if 0 <= int(self.e_overdraft1.get()
                                    ) <= 1000 and 0 <= float(
                                        self.e_intrest1.get()) <= 5.00:
                            if self.e_acc_type1.get(
                            ) == "Current Account" and self.e_intrest1.get(
                            ) != "0":
                                tkinter.messagebox.showinfo(
                                    "Account Management:",
                                    f'The rate of Intrest for {self.e_acc_type1.get()} is 0%. Please enter your details again.'
                                )
                                self.e_intrest1.delete(0, 'end')
                                break
                            elif self.e_acc_type1.get(
                            ) == "Saving Account" and self.e_overdraft1.get(
                            ) != "0":
                                tkinter.messagebox.showinfo(
                                    "Account Management:",
                                    f'The Overdraft limit for {self.e_acc_type1.get()} is £0. Please enter your details again.'
                                )
                                self.e_overdraft1.delete(0, 'end')
                                break
                            elif self.e_acc_type1.get(
                            ) == "Saving Account" and self.e_overdraft1.get(
                            ) == "0" or self.e_acc_type1.get(
                            ) == "Current Account" and self.e_intrest1.get(
                            ) == "0":
                                if int(self.e_Balance1.get()) >= 0:
                                    self.ws.cell(row, 9).value = float(
                                        self.e_intrest1.get())
                                    self.ws.cell(row, 8).value = int(
                                        self.e_overdraft1.get())
                                    self.ws.cell(
                                        row, 1).value = self.e_username1.get()
                                    self.ws.cell(
                                        row, 2).value = self.e_password1.get()
                                    self.ws.cell(
                                        row,
                                        3).value = self.e_user_type1.get()
                                    self.ws.cell(row, 4).value = int(
                                        self.e_Balance1.get())
                                    self.ws.cell(
                                        row, 5).value = self.e_Address1.get()
                                    self.ws.cell(
                                        row, 6).value = self.e_acc_type1.get()
                                    self.ws.cell(row, 7).value = int(
                                        self.e_acc_no1.get())
                                    tkinter.messagebox.showinfo(
                                        "Account Management:",
                                        f'You added {self.e_acc_type1.get()} with username {self.e_username1.get()}.'
                                    )
                                    self.save_user_file()
                                    self.add_reset_form()
                                    self.mw.destroy()
                                    break
                                else:
                                    tkinter.messagebox.showinfo(
                                        "Account Management:",
                                        f'For opening an Account, the Balance can not be negative.'
                                    )
                                    break

                            elif self.e_acc_type1.get(
                            ) != "Saving Account" or self.e_acc_type1.get(
                            ) != "Current Account":
                                tkinter.messagebox.showinfo(
                                    "Account Management:",
                                    f'The type of Account, you entered does not exist. We have Current Account or Saving Account as type of Accounts. Please enter again.'
                                )
                                self.e_acc_type1.delete(0, 'end')
                                break

                            elif self.e_user_type1.get(
                            ) != "Admin" or self.e_user_type1.get(
                            ) != "Customer":
                                tkinter.messagebox.showinfo(
                                    "Account Management:",
                                    f'The type of User, you entered does not exist. We have Admin or Customer as types of User. Please enter again.'
                                )
                                self.e_user_type1.delete(0, 'end')
                                break
                        else:
                            tkinter.messagebox.showinfo(
                                "Account Management:",
                                f'The Overdraft limit or Intrest rate is out of range. we have Overdraft limit of £1000 and max-Intrest rate of 5.00%. Please enter again.'
                            )
                            self.e_overdraft1.delete(0, 'end')
                            self.e_intrest1.delete(0, 'end')
                            break
        except ValueError:
            tkinter.messagebox.showinfo(
                "TRANSACTION:",
                f'Every entery box is neccassary and Please enter proper value as per required.'
            )

    def delete_account(self):
        self.user_file()
        for row in range(1, 20):
            if self.d_username1.get() == self.ws.cell(
                    row, 1).value and self.d_password1.get() == self.ws.cell(
                        row, 2).value and self.d_acc_no1.get() == str(
                            self.ws.cell(row, 7).value):
                for i in range(1, 20):
                    if i != row:
                        if self.d_username1.get() == self.ws.cell(
                                i, 1).value and self.d_password1.get(
                                ) == self.ws.cell(i, 2).value:
                            self.ws.delete_rows(row, 1)
                            tkinter.messagebox.showinfo(
                                "Account Management:",
                                f'You have deleted account of {self.d_username1.get()} with account number {self.d_acc_no1.get()}.'
                            )
                            self.save_user_file()
                            self.delete_reset_form()
                            self.mw.destroy()
                            break
                else:
                    tkinter.messagebox.showinfo(
                        "Account Management:",
                        f'The username {self.d_username1.get()} with account number {self.d_acc_no1.get()} has the last account in the bank. As per policy user can not delete last account.'
                    )
                    self.delete_reset_form()
                break

        else:
            tkinter.messagebox.showinfo(
                "Account Management:",
                f'The entered details are incorrect. Please enter again.')
            self.delete_reset_form()

    def add_reset_form(self):
        self.e_username1.delete(0, 'end')
        self.e_password1.delete(0, 'end')
        self.e_user_type1.delete(0, 'end')
        self.e_Balance1.delete(0, 'end')
        self.e_Address1.delete(0, 'end')
        self.e_acc_type1.delete(0, 'end')
        self.e_acc_no1.delete(0, 'end')
        self.e_overdraft1.delete(0, 'end')
        self.e_intrest1.delete(0, 'end')

    def delete_reset_form(self):
        self.d_username1.delete(0, 'end')
        self.d_password1.delete(0, 'end')
        self.d_acc_no1.delete(0, 'end')

    def Money_transfer_reset(self):
        self.ask_from_username.delete(0, 'end')
        self.ask_from_account.delete(0, 'end')
        self.ask_to_username.delete(0, 'end')
        self.ask_to_account.delete(0, 'end')
        self.ask_amount.delete(0, 'end')


class Customer(User):
    def __init__(self, name, password, uType, address):
        User.__init__(self, name, password, uType)
        self.cust_menu = ("Please select an option:", "  1 - View account",
                          "  2 - View summary", "  3 - Quit")
        self.address = address
        self.cust_operation = ("Please select an option:", "  1 - Deposit",
                               "  2 - Withdraw", "  3 - Go back")

    def display_Cus_menu(self):
        self.user_file()
        for i in self.cust_menu:
            print(i)
        self.ask_input1 = input("Enter a number to select your option:")
        if self.ask_input1 == "1":
            self.display_account()
        elif self.ask_input1 == "2":
            self.dis_cus_account_summary()
        elif self.ask_input1 == "3":
            print("Thank you for using V - Banking\n----- Bye -----")
        else:
            print("----- INVALID INPUT -----\n...please enter again...")
            self.display_Cus_menu()

    def display_account(self):
        self.no_of_acc = 1
        print("---- Account list ----")
        for row in range(1, 20):
            if self.name == self.ws.cell(
                    row, 1).value and self.password == self.ws.cell(row,
                                                                    2).value:
                print(
                    f'{self.no_of_acc} - {self.ws.cell(row, 6).value}: £{self.ws.cell(row, 4).value}'
                )
                self.no_of_acc += 1
        print(f'{self.no_of_acc} - Go back')
        self.dis_selected_acc()

    def dis_cus_account_summary(self):
        self.no_of_acc = 1
        self.total_balance = 0
        for row in range(1, 20):
            if self.name == self.ws.cell(
                    row, 1).value and self.password == self.ws.cell(row,
                                                                    2).value:
                self.total_balance += self.ws.cell(row, 4).value
                self.no_of_acc += 1
                continue
        print(f'Total accounts: {self.no_of_acc-1}')
        print(f'Total balance: £{self.total_balance}')
        print(f'address: {self.address}')

    def dis_selected_acc(self):
        self.ask_for_acc = int(input("Enter a number to select your option:"))
        try:
            if self.ask_for_acc in range(1, self.no_of_acc):
                for row in range(1, 20):
                    if self.name == self.ws.cell(row, 1).value:
                        if self.ask_for_acc == self.ws.cell(
                                row, 7).value and self.ws.cell(
                                    row, 6).value == "Saving Account":
                            print(
                                f'You selected {self.ask_for_acc} - {self.ws.cell(row, 6).value}: £{self.ws.cell(row, 4).value}.'
                            )
                            self.customer_ope()
                        elif self.ask_for_acc == self.ws.cell(
                                row, 7).value and self.ws.cell(
                                    row, 6).value == "Current Account":
                            print(
                                f'You selected {self.ask_for_acc} - {self.ws.cell(row, 6).value}: £{self.ws.cell(row, 4).value}.'
                            )
                            self.customer_ope()

            elif self.ask_for_acc == self.no_of_acc:
                self.display_Cus_menu()

            else:
                print("----- INVALID INPUT -----\n...please enter again...")
                self.display_account()

        except ValueError as error:
            print(str(error))

    def customer_ope(self):
        for i in self.cust_operation:
            print(i)
        self.ask_for_ope = input("Enter a number to select your option:")
        if self.ask_for_ope == "1":
            self.deposite_amount()
            self.display_account()
        elif self.ask_for_ope == "2":
            self.withdraw_amount()
            self.display_account()
        elif self.ask_for_ope == "3":
            self.display_account()
        else:
            print("----- INVALID INPUT -----\n...please enter again...")
            self.customer_ope()

    def deposite_amount(self):
        self.ask_for_depo = int(input("Enter amount to deposit: £"))
        try:
            if self.ask_for_depo > 0:
                for row in range(1, 20):
                    if self.name in self.ws.cell(row, 1).value:
                        if self.ask_for_acc == self.ws.cell(
                                row, 7
                        ).value and self.ws.cell(
                                row, 6
                        ).value == "Saving Account" or self.ask_for_acc == self.ws.cell(
                                row, 7).value and self.ws.cell(
                                    row, 6).value == "Current Account":
                            self.ws.cell(row, 4).value += self.ask_for_depo
                            self.ws.cell(row, 10).value += self.ask_for_depo
                            print(
                                f'You have deposited £{self.ask_for_depo} to your {self.ws.cell(row,6).value}.'
                            )
                            self.save_user_file()
                            break
            else:
                print(
                    "The deposited amount can not 0 or less than 0. Please enter again."
                )
                self.deposite_amount()
        except ValueError as error:
            print(str(error))

    def withdraw_amount(self):
        self.ask_for_with = int(input("Enter amount to withdraw: £"))
        try:
            if self.ask_for_with > 0:
                for row in range(1, 20):
                    if self.name == self.ws.cell(
                            row, 1).value and self.ask_for_acc == self.ws.cell(
                                row, 7).value:
                        if 0 <= self.ws.cell(row, 8).value <= 1000:
                            if self.ws.cell(row, 4).value > -abs(
                                    self.ws.cell(row,
                                                 8).value) or self.ws.cell(
                                                     row, 4).value > 0:
                                if self.ws.cell(
                                        row,
                                        4).value - self.ask_for_with < -abs(
                                            self.ws.cell(row, 8).value
                                        ) and self.ws.cell(
                                            row, 6).value == "Current Account":
                                    print(
                                        f'The ammount is large compared to balance.\nYou can withdraw only £{self.ws.cell(row,4).value+self.ws.cell(row,8).value}.'
                                    )
                                    break
                                elif self.ws.cell(
                                        row, 4
                                ).value - self.ask_for_with < 0 and self.ws.cell(
                                        row, 6).value == "Saving Account":
                                    print(
                                        f'The ammount is large compared to balance.\nYou can withdraw only £{self.ws.cell(row,4).value+self.ws.cell(row,8).value}.'
                                    )
                                    break

                                else:
                                    self.ws.cell(row,
                                                 4).value -= self.ask_for_with
                                    self.ws.cell(row,
                                                 11).value += self.ask_for_with
                                    print(
                                        f'You have withdrawn £{self.ask_for_with} from your {self.ws.cell(row,6).value}.'
                                    )
                                    self.save_user_file()
                                    break
                            elif self.ws.cell(row, 4).value == -abs(
                                    self.ws.cell(row,
                                                 8).value) or self.ws.cell(
                                                     row, 4).value == 0:
                                print(
                                    f'This account balance is {self.ws.cell(row, 4).value}.\nYou can not withdraw money.'
                                )
                                break

            else:
                print(
                    "The deposited amount can not 0 or less than 0. Please enter again."
                )
                self.deposite_amount()
        except ValueError as error:
            print(str(error))
